import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import FastAPI, Depends, HTTPException, Request
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from typing import List
from . import models, schemas, crud, database

# Cria as tabelas
models.Base.metadata.create_all(bind=database.engine)

app = FastAPI(
    title="Stock Flow API Pro",
    description="Sistema Inteligente de Controle de Estoque",
    version="2.6.0"
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Estáticos e Templates
app.mount("/static", StaticFiles(directory="static"), name="static")
templates = Jinja2Templates(directory="templates")

# --- FRONTEND ---
@app.get("/", response_class=HTMLResponse, include_in_schema=False)
async def view_dashboard(request: Request):
    return templates.TemplateResponse("index.html", {"request": request, "page": "dashboard"})

@app.get("/products-ui", response_class=HTMLResponse, include_in_schema=False)
async def view_products(request: Request):
    return templates.TemplateResponse("products.html", {"request": request, "page": "products"})

@app.get("/movements-ui", response_class=HTMLResponse, include_in_schema=False)
async def view_movements(request: Request):
    return templates.TemplateResponse("movements.html", {"request": request, "page": "movements"})

# --- API ---
def get_db():
    db = database.SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Produtos
@app.post("/products/", response_model=schemas.ProductResponse, tags=["Produtos"])
def criar_produto(product: schemas.ProductCreate, db: Session = Depends(get_db)):
    return crud.create_product(db=db, product=product)

@app.get("/products/", response_model=List[schemas.ProductResponse], tags=["Produtos"])
def listar_produtos(db: Session = Depends(get_db)):
    return crud.get_products(db)

@app.put("/products/{product_id}", response_model=schemas.ProductResponse, tags=["Produtos"])
def atualizar_produto(product_id: int, product: schemas.ProductUpdate, db: Session = Depends(get_db)):
    db_product = crud.update_product(db, product_id, product)
    if db_product is None:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    return db_product

@app.delete("/products/{product_id}", tags=["Produtos"])
def deletar_produto(product_id: int, db: Session = Depends(get_db)):
    success = crud.delete_product(db, product_id)
    if not success:
        raise HTTPException(status_code=404, detail="Produto não encontrado")
    return {"detail": "Produto deletado com sucesso"}

# NOVO: Histórico do Produto Específico
@app.get("/products/{product_id}/movements", response_model=List[schemas.MovementResponse], tags=["Produtos"])
def historico_produto(product_id: int, db: Session = Depends(get_db)):
    return crud.get_movements_by_product(db, product_id)

# Movimentações
@app.post("/movements/", response_model=schemas.MovementResponse, tags=["Movimentação"])
def registrar_movimento(movement: schemas.MovementCreate, db: Session = Depends(get_db)):
    return crud.create_movement(db=db, movement=movement)

@app.get("/movements/recent", response_model=List[schemas.MovementResponse], tags=["Movimentação"])
def movimentacoes_recentes(db: Session = Depends(get_db)):
    return crud.get_recent_movements(db)

# Relatórios
@app.get("/reports/low-stock", response_model=List[schemas.ProductResponse], tags=["Relatórios"])
def relatorio_estoque_baixo(db: Session = Depends(get_db)):
    return crud.get_low_stock(db)

@app.get("/reports/export", tags=["Relatórios"])
def exportar_relatorio(db: Session = Depends(get_db)):
    products = crud.get_products(db)
    wb = Workbook()
    ws = wb.active
    ws.title = "Estoque Atual"

    headers = ["ID", "Produto", "Descrição", "Preço (R$)", "Qtd Atual", "Qtd Mínima", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for p in products:
        status = "OK" if p.quantity > p.min_stock else "REPOR"
        ws.append([p.id, p.name, p.description, p.price / 100, p.quantity, p.min_stock, status])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[3].number_format = 'R$ #,##0.00'
        if row[6].value == "REPOR":
            row[6].font = Font(color="DC3545", bold=True)
        else:
            row[6].font = Font(color="198754", bold=True)

    for col in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = length + 5

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    response = StreamingResponse(
        iter([stream.getvalue()]), 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response.headers["Content-Disposition"] = "attachment; filename=Relatorio_Estoque.xlsx"
    return response